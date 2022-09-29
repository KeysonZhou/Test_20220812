import openpyxl

class TestData:
    def __init__(self,file,sheetname):
        self.file=file
        self.wb=openpyxl.load_workbook(self.file)
        self.table=self.wb[sheetname]
        self.mrow=self.table.max_row
        self.mcol=self.table.max_column
        self.wb.close()

    def read_xlsx(self):
        list_data=[]
        for row in range(2,self.mrow+1):
            list_row=[]
            for col in range(1,self.mcol+1):
                list_row.append(self.table.cell(row,col).value)
            list_data.append(list_row)
        return list_data